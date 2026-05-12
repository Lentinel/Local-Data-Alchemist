import os
import csv
import base64
import io
from pathlib import Path
from collections import Counter
from datetime import date, datetime, timedelta

from models.schemas import FileInfo
from utils.helpers import classify_file
from utils.constants import (
    TEXT_EXTENSIONS,
    IMAGE_EXTENSIONS,
    CATEGORY_LABELS,
)
from utils.security import resolve_inside_target, to_target_relative
from utils.storage import list_history

# PDF 解析（延迟导入，避免未安装时崩溃）
_pdfplumber_available = False
try:
    import pdfplumber
    _pdfplumber_available = True
except ImportError:
    pass

# Excel 解析（延迟导入）
_openpyxl_available = False
try:
    import openpyxl
    _openpyxl_available = True
except ImportError:
    pass


def scan_target_files(target_dir: Path) -> list[dict]:
    files_info = []
    ignored_names = {"snapshot.json"}
    ignored_dirs = {".alchemy_trash"}

    for root, dirs, files in os.walk(target_dir):
        dirs[:] = [directory for directory in dirs if directory not in ignored_dirs]
        for filename in files:
            if filename in ignored_names:
                continue

            file_path = Path(root) / filename
            file_rel_path = file_path.relative_to(target_dir).as_posix()
            extension = file_path.suffix.lower()
            files_info.append(
                {
                    "name": filename,
                    "path": file_rel_path,
                    "extension": extension,
                    "category": classify_file(extension, filename),
                    "size": file_path.stat().st_size,
                }
            )

    return files_info


def _extract_pdf_text(file_path: Path, max_chars: int = 2000) -> str:
    """从 PDF 文件中提取文本内容。
    
    Args:
        file_path: PDF 文件路径
        max_chars: 最大提取字符数
        
    Returns:
        提取的文本内容，失败时返回错误信息
    """
    if not _pdfplumber_available:
        return "[PDF 解析库未安装，请运行: pip install pdfplumber]"
    
    try:
        text_parts = []
        with pdfplumber.open(str(file_path)) as pdf:
            for i, page in enumerate(pdf.pages[:5]):  # 最多读取前5页
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text.strip())
                if sum(len(t) for t in text_parts) >= max_chars:
                    break
            
            # 提取表格数据
            tables_text = []
            for page in pdf.pages[:3]:  # 前3页的表格
                tables = page.extract_tables()
                for table in tables[:2]:  # 每页最多2个表格
                    for row in table[:10]:  # 每个表格最多10行
                        if row:
                            tables_text.append(" | ".join(str(cell or "") for cell in row))
        
        result = "\n".join(text_parts)
        if tables_text:
            result += "\n\n[表格数据]\n" + "\n".join(tables_text)
        
        return result[:max_chars] if result else "[PDF 文件无文本内容（可能是扫描件）]"
    except Exception as e:
        return f"[PDF 解析失败: {str(e)}]"


def _extract_excel_info(file_path: Path, max_rows: int = 20) -> str:
    """从 Excel 文件中提取结构化信息。
    
    Args:
        file_path: Excel 文件路径
        max_rows: 最大读取行数
        
    Returns:
        提取的结构化信息
    """
    if not _openpyxl_available:
        return "[Excel 解析库未安装，请运行: pip install openpyxl]"
    
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        info_parts = []
        
        for sheet_name in wb.sheetnames[:3]:  # 最多3个sheet
            ws = wb[sheet_name]
            info_parts.append(f"[Sheet: {sheet_name}]")
            
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    info_parts.append(f"... (共 {ws.max_row} 行，仅显示前 {max_rows} 行)")
                    break
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cells):  # 跳过空行
                    rows.append(" | ".join(cells))
            
            info_parts.extend(rows)
            info_parts.append("")
        
        wb.close()
        return "\n".join(info_parts) if info_parts else "[Excel 文件为空]"
    except Exception as e:
        return f"[Excel 解析失败: {str(e)}]"


def _extract_csv_info(file_path: Path, max_rows: int = 20) -> str:
    """从 CSV 文件中提取结构化信息。
    
    Args:
        file_path: CSV 文件路径
        max_rows: 最大读取行数
        
    Returns:
        提取的结构化信息
    """
    try:
        with file_path.open("r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            rows = []
            for i, row in enumerate(reader):
                if i >= max_rows:
                    rows.append(f"... (仅显示前 {max_rows} 行)")
                    break
                rows.append(" | ".join(row))
            return "\n".join(rows) if rows else "[CSV 文件为空]"
    except Exception as e:
        return f"[CSV 解析失败: {str(e)}]"


def peek_file_content(file_path: Path, max_bytes: int = 512) -> str:
    """读取文件内容摘要。
    
    支持文本文件、PDF、Excel、CSV 等格式。
    
    Args:
        file_path: 文件路径
        max_bytes: 文本文件最大读取字节数
        
    Returns:
        文件内容摘要
    """
    extension = file_path.suffix.lower()
    
    # PDF 文件
    if extension == ".pdf":
        return _extract_pdf_text(file_path, max_chars=max_bytes * 4)
    
    # Excel 文件
    if extension in {".xlsx", ".xls"}:
        return _extract_excel_info(file_path)
    
    # CSV 文件
    if extension == ".csv":
        return _extract_csv_info(file_path)
    
    # 文本文件
    if extension in TEXT_EXTENSIONS:
        try:
            with file_path.open("rb") as file:
                return file.read(max_bytes).decode("utf-8", errors="ignore").strip()
        except OSError:
            return "[Unreadable File]"
    
    # 图片文件 - 使用多模态 LLM 分析
    if extension in IMAGE_EXTENSIONS:
        try:
            from services.llm_service import analyze_image_with_llm
            return analyze_image_with_llm(file_path)
        except Exception:
            return "[图片文件]"
    
    return "[Binary File]"


def get_file_preview(file_path: Path, max_bytes: int = 10240, max_image_bytes: int = 10 * 1024 * 1024) -> dict:
    extension = file_path.suffix.lower()
    file_name = file_path.name
    
    try:
        file_size = file_path.stat().st_size
    except (OSError, Exception):
        file_size = 0

    preview = {
        "name": file_name,
        "path": str(file_path),
        "extension": extension,
        "size": file_size,
        "type": "unknown",
        "content": None,
        "truncated": False,
    }

    if extension in TEXT_EXTENSIONS:
        preview["type"] = "text"
        try:
            with file_path.open("rb") as f:
                content_bytes = f.read(max_bytes + 1)
                if len(content_bytes) > max_bytes:
                    preview["truncated"] = True
                    content_bytes = content_bytes[:max_bytes]
                preview["content"] = content_bytes.decode("utf-8", errors="ignore")
        except (OSError, Exception):
            preview["content"] = "[无法读取文件]"
        return preview

    if extension in IMAGE_EXTENSIONS:
        preview["type"] = "image"
        
        if file_size > max_image_bytes:
            preview["content"] = f"[图片文件过大，超过 {max_image_bytes // 1024 // 1024}MB，不支持预览]"
            return preview
        
        try:
            with file_path.open("rb") as f:
                preview["content"] = base64.b64encode(f.read()).decode("utf-8")
        except (OSError, Exception):
            preview["content"] = None
        return preview

    if extension == ".pdf":
        preview["type"] = "pdf"
        preview["content"] = _extract_pdf_text(file_path, max_chars=max_bytes * 2)
        preview["page_count"] = None
        if _pdfplumber_available:
            try:
                with pdfplumber.open(str(file_path)) as pdf:
                    preview["page_count"] = len(pdf.pages)
            except Exception:
                pass
        return preview

    if extension in {".xlsx", ".xls"}:
        preview["type"] = "spreadsheet"
        preview["content"] = _extract_excel_info(file_path)
        return preview

    if extension == ".csv":
        preview["type"] = "spreadsheet"
        preview["content"] = _extract_csv_info(file_path)
        return preview

    preview["type"] = "binary"
    preview["content"] = "[二进制文件，不支持预览]"
    return preview


def build_directory_tree(target_dir: Path, plan: list[dict] = None) -> dict:
    """构建目录树结构，可选包含计划中的目标路径。
    
    Args:
        target_dir: 目标目录路径
        plan: 可选的执行计划，用于预览执行后的目录结构
        
    Returns:
        目录树结构字典
    """
    tree = {
        "name": target_dir.name or str(target_dir),
        "path": ".",
        "type": "directory",
        "children": [],
        "file_count": 0,
        "total_size": 0,
    }
    
    # 扫描当前目录结构
    def scan_dir(current_path: Path, relative_path: str = ".", node: dict = None):
        if node is None:
            node = tree
        
        try:
            entries = sorted(current_path.iterdir(), key=lambda e: (not e.is_dir(), e.name.lower()))
        except PermissionError:
            return
        
        for entry in entries:
            # 跳过隐藏文件和系统文件
            if entry.name.startswith('.') or entry.name in {'snapshot.json'}:
                continue
            
            entry_relative = f"{relative_path}/{entry.name}" if relative_path != "." else entry.name
            
            if entry.is_dir():
                if entry.name == '.alchemy_trash':
                    continue
                dir_node = {
                    "name": entry.name,
                    "path": entry_relative,
                    "type": "directory",
                    "children": [],
                    "file_count": 0,
                    "total_size": 0,
                }
                node["children"].append(dir_node)
                scan_dir(entry, entry_relative, dir_node)
                node["file_count"] += dir_node["file_count"]
                node["total_size"] += dir_node["total_size"]
            else:
                try:
                    file_size = entry.stat().st_size
                except OSError:
                    file_size = 0
                
                file_node = {
                    "name": entry.name,
                    "path": entry_relative,
                    "type": "file",
                    "extension": entry.suffix.lower(),
                    "size": file_size,
                    "category": classify_file(entry.suffix.lower()),
                }
                node["children"].append(file_node)
                node["file_count"] += 1
                node["total_size"] += file_size
    
    scan_dir(target_dir)
    
    # 如果有计划，添加计划中的目标路径预览
    if plan:
        planned_paths = set()
        for item in plan:
            if item.get("target_path") and item.get("action") in {"move", "rename_and_move"}:
                planned_paths.add(item["target_path"])
        
        if planned_paths:
            tree["planned_structure"] = _build_planned_tree(planned_paths)
    
    return tree


def _build_planned_tree(planned_paths: set) -> dict:
    """从计划路径构建预览目录树。"""
    root = {
        "name": "执行后预览",
        "path": ".",
        "type": "directory",
        "children": [],
    }
    
    for path in sorted(planned_paths):
        parts = path.replace("\\", "/").split("/")
        current = root
        
        for i, part in enumerate(parts):
            if i == len(parts) - 1:
                # 文件
                current["children"].append({
                    "name": part,
                    "path": path,
                    "type": "file",
                    "is_planned": True,
                })
            else:
                # 目录
                found = False
                for child in current["children"]:
                    if child["name"] == part and child["type"] == "directory":
                        current = child
                        found = True
                        break
                
                if not found:
                    new_dir = {
                        "name": part,
                        "path": "/".join(parts[:i+1]),
                        "type": "directory",
                        "children": [],
                        "is_planned": True,
                    }
                    current["children"].append(new_dir)
                    current = new_dir
    
    return root


def build_analysis(files_info: list[dict]) -> dict:
    category_counter = Counter(file["category"] for file in files_info)
    extension_counter = Counter(file["extension"] or "unknown" for file in files_info)

    recommendations = []
    if category_counter["logs"]:
        recommendations.append("检测到日志文件，可优先进行错误码、时间线和异常频率统计。")
    if category_counter["images"]:
        recommendations.append("检测到图片文件，下一步可接入 OCR 或多模态模型提取票据/截图信息。")
    if category_counter["documents"]:
        recommendations.append("检测到文档或表格，可进入结构化字段抽取与汇总。")
    if category_counter["unknown"]:
        recommendations.append("存在未知类型文件，建议保留原路径并交给后续 AI 分类器复判。")
    if not recommendations:
        recommendations.append("文件清单已生成，可继续接入 LLM 分析流水线。")

    return {
        "mode": "local-rule-engine",
        "total_files": len(files_info),
        "categories": [
            {
                "key": key,
                "label": CATEGORY_LABELS[key],
                "count": category_counter[key],
            }
            for key in CATEGORY_LABELS
            if category_counter[key]
        ],
        "extensions": [
            {"extension": extension, "count": count}
            for extension, count in extension_counter.most_common()
        ],
        "recommendations": recommendations,
    }


def build_file_snippets(target_dir: Path, files: list[FileInfo]) -> list[dict]:
    snippets = []
    for file in files:
        file_path = resolve_inside_target(target_dir, file.path)
        snippets.append(
            {
                "filename": file.path,
                "extension": file.extension,
                "category": file.category,
                "size": file.size,
                "snippet": peek_file_content(file_path),
            }
        )
    return snippets


def get_file_modification_time(file_path: Path) -> date | None:
    try:
        timestamp = file_path.stat().st_mtime
        return date.fromtimestamp(timestamp)
    except (OSError, Exception):
        return None


def calculate_dashboard_stats(
    target_dir: Path,
    files_info: list[dict],
    include_history: bool = True
) -> dict:
    total_files = len(files_info)
    total_size = sum(f.get("size", 0) for f in files_info)
    
    category_stats = {
        "images": {"count": 0, "size": 0},
        "documents": {"count": 0, "size": 0},
        "archives": {"count": 0, "size": 0},
        "code": {"count": 0, "size": 0},
        "logs": {"count": 0, "size": 0},
        "unknown": {"count": 0, "size": 0},
    }
    
    extension_counts = {}
    
    size_buckets = {
        "tiny": {"count": 0, "size": 0, "label": "< 100KB"},
        "small": {"count": 0, "size": 0, "label": "100KB - 1MB"},
        "medium": {"count": 0, "size": 0, "label": "1MB - 10MB"},
        "large": {"count": 0, "size": 0, "label": "10MB - 100MB"},
        "huge": {"count": 0, "size": 0, "label": "> 100MB"},
    }
    
    today = date.today()
    week_dates = []
    for i in range(6, -1, -1):
        week_dates.append(today - timedelta(days=i))
    
    weekly_stats = {}
    for d in week_dates:
        weekly_stats[d.isoformat()] = {"count": 0, "size": 0, "date": d.isoformat()}
    
    directory_tree = {
        "name": ".",
        "path": ".",
        "count": 0,
        "size": 0,
        "children": {}
    }
    
    all_files = []
    
    for file_info in files_info:
        category = file_info.get("category", "unknown")
        size = file_info.get("size", 0)
        extension = file_info.get("extension", "").lower()
        file_path_str = file_info.get("path", "")
        file_name = file_info.get("name", file_path_str)
        
        all_files.append({
            "name": file_name,
            "path": file_path_str,
            "size": size,
            "extension": extension,
            "category": category
        })
        
        if category in category_stats:
            category_stats[category]["count"] += 1
            category_stats[category]["size"] += size
        
        if extension:
            if extension not in extension_counts:
                extension_counts[extension] = {"count": 0, "size": 0}
            extension_counts[extension]["count"] += 1
            extension_counts[extension]["size"] += size
        
        if size < 100 * 1024:
            size_buckets["tiny"]["count"] += 1
            size_buckets["tiny"]["size"] += size
        elif size < 1024 * 1024:
            size_buckets["small"]["count"] += 1
            size_buckets["small"]["size"] += size
        elif size < 10 * 1024 * 1024:
            size_buckets["medium"]["count"] += 1
            size_buckets["medium"]["size"] += size
        elif size < 100 * 1024 * 1024:
            size_buckets["large"]["count"] += 1
            size_buckets["large"]["size"] += size
        else:
            size_buckets["huge"]["count"] += 1
            size_buckets["huge"]["size"] += size
        
        file_path = resolve_inside_target(target_dir, file_path_str)
        if file_path.exists():
            mod_date = get_file_modification_time(file_path)
            if mod_date:
                date_str = mod_date.isoformat()
                if date_str in weekly_stats:
                    weekly_stats[date_str]["count"] += 1
                    weekly_stats[date_str]["size"] += size
        
        parent_dir = Path(file_path_str).parent
        current = directory_tree
        current["count"] += 1
        current["size"] += size
        
        if parent_dir != Path("."):
            parts = parent_dir.as_posix().split("/")
            for part in parts:
                if part and part != ".":
                    if part not in current["children"]:
                        current["children"][part] = {
                            "name": part,
                            "path": part,
                            "count": 0,
                            "size": 0,
                            "children": {}
                        }
                    current = current["children"][part]
                    current["count"] += 1
                    current["size"] += size
    
    def flatten_tree(node, parent_path=""):
        result = []
        current_path = parent_path + "/" + node["name"] if parent_path else node["name"]
        node["full_path"] = current_path
        result.append(node)
        
        for child in node["children"].values():
            result.extend(flatten_tree(child, current_path))
        
        return result
    
    flattened_dirs = flatten_tree(directory_tree)
    directory_list = sorted(
        [d for d in flattened_dirs if d["name"] != "."],
        key=lambda x: x["count"],
        reverse=True
    )[:20]
    
    sorted_files = sorted(all_files, key=lambda x: x["size"], reverse=True)
    largest_files = sorted_files[:10]
    
    top_extensions = sorted(
        extension_counts.items(),
        key=lambda x: x[1]["count"],
        reverse=True
    )[:15]
    
    history_stats = None
    if include_history:
        history_list = list_history(target_dir)
        history_stats = {
            "total_operations": len(history_list),
            "by_type": {},
            "last_7_days": 0,
        }
        
        for item in history_list:
            op_type = item.get("type", "unknown")
            if op_type not in history_stats["by_type"]:
                history_stats["by_type"][op_type] = 0
            history_stats["by_type"][op_type] += 1
            
            created_at = item.get("created_at")
            if created_at:
                try:
                    if isinstance(created_at, str):
                        if "T" in created_at:
                            item_date = datetime.fromisoformat(created_at).date()
                        else:
                            item_date = date.fromisoformat(created_at)
                    else:
                        item_date = created_at
                    
                    if (today - item_date).days <= 7:
                        history_stats["last_7_days"] += 1
                except (ValueError, TypeError):
                    pass
    
    def tree_to_dict(node):
        return {
            "name": node["name"],
            "path": node["path"],
            "count": node["count"],
            "size": node["size"],
            "children": {k: tree_to_dict(v) for k, v in node["children"].items()}
        }
    
    return {
        "overview": {
            "total_files": total_files,
            "total_size": total_size,
            "categories": category_stats,
        },
        "size_distribution": size_buckets,
        "extension_distribution": {
            ext: data for ext, data in top_extensions
        },
        "weekly_activity": list(weekly_stats.values()),
        "history_stats": history_stats,
        "directory_tree": tree_to_dict(directory_tree),
        "directory_list": directory_list,
        "largest_files": largest_files,
        "all_files_count": len(all_files),
        "categories_detail": {
            k: v for k, v in category_stats.items() if v["count"] > 0
        }
    }


def preview_plan_impl(
    target_dir: Path,
    plan: list
) -> dict:
    from utils.constants import ALLOWED_ACTIONS
    
    summary = {
        "total_actions": len(plan),
        "by_action": {},
        "move_actions": [],
        "delete_actions": [],
        "keep_actions": [],
        "conflicts": [],
        "warnings": [],
        "new_folders": set(),
    }

    target_paths_map = {}
    source_paths_map = {}

    for i, item in enumerate(plan):
        action = item.action
        source_file = item.file

        if action not in summary["by_action"]:
            summary["by_action"][action] = 0
        summary["by_action"][action] += 1

        source_path = resolve_inside_target(target_dir, source_file)
        source_exists = source_path.exists()

        if not source_exists:
            summary["warnings"].append({
                "index": i,
                "action": action,
                "file": source_file,
                "message": "源文件不存在"
            })

        if action in {"move", "rename_and_move"}:
            target_file = item.target_path
            target_path = resolve_inside_target(target_dir, target_file)
            
            summary["move_actions"].append({
                "index": i,
                "source": source_file,
                "source_name": Path(source_file).name,
                "target": target_file,
                "target_name": Path(target_file).name,
                "source_exists": source_exists,
            })

            target_parent = Path(target_file).parent
            if str(target_parent) != "." and str(target_parent) != "":
                summary["new_folders"].add(str(target_parent))

            if target_path.exists() and target_path != source_path:
                summary["conflicts"].append({
                    "index": i,
                    "type": "target_exists",
                    "source": source_file,
                    "target": target_file,
                    "message": "目标路径已存在"
                })

            normalized_target = str(target_path.resolve())
            if normalized_target in target_paths_map:
                summary["conflicts"].append({
                    "index": i,
                    "type": "duplicate_target",
                    "source": source_file,
                    "target": target_file,
                    "other_source": target_paths_map[normalized_target]["source"],
                    "message": "多个文件将移动到同一个目标"
                })
            target_paths_map[normalized_target] = {
                "source": source_file,
                "target": target_file,
                "index": i,
            }

        elif action == "delete":
            summary["delete_actions"].append({
                "index": i,
                "file": source_file,
                "file_name": Path(source_file).name,
                "source_exists": source_exists,
            })

        elif action == "keep":
            summary["keep_actions"].append({
                "index": i,
                "file": source_file,
                "file_name": Path(source_file).name,
                "source_exists": source_exists,
            })

    move_count = len(summary["move_actions"])
    delete_count = len(summary["delete_actions"])
    keep_count = len(summary["keep_actions"])

    has_conflicts = len(summary["conflicts"]) > 0
    has_warnings = len(summary["warnings"]) > 0
    needs_confirmation = has_conflicts or delete_count > 0

    safety_level = "safe"
    safety_reasons = []

    if has_conflicts:
        safety_level = "danger"
        safety_reasons.append(f"检测到 {len(summary['conflicts'])} 个冲突")
    
    if delete_count > 0:
        safety_level = "warning" if safety_level == "safe" else safety_level
        safety_reasons.append(f"将删除 {delete_count} 个文件")
    
    if move_count > 10:
        safety_level = "warning" if safety_level == "safe" else safety_level
        safety_reasons.append(f"涉及 {move_count} 个文件操作")

    return {
        "status": "success",
        "summary": {
            "total_actions": summary["total_actions"],
            "move_count": move_count,
            "delete_count": delete_count,
            "keep_count": keep_count,
            "by_action": summary["by_action"],
            "new_folders_count": len(summary["new_folders"]),
            "new_folders": list(summary["new_folders"]),
        },
        "conflicts": summary["conflicts"],
        "warnings": summary["warnings"],
        "move_actions": summary["move_actions"],
        "delete_actions": summary["delete_actions"],
        "keep_actions": summary["keep_actions"],
        "safety_level": safety_level,
        "safety_reasons": safety_reasons,
        "has_conflicts": has_conflicts,
        "has_warnings": has_warnings,
        "needs_confirmation": needs_confirmation,
    }


def preview_file_impl(target_dir: Path, relative_file_path: str) -> dict:
    """预览文件内容。
    
    Args:
        target_dir: 目标目录路径
        relative_file_path: 文件相对路径
        
    Returns:
        包含预览信息的字典
    """
    from fastapi import HTTPException
    from utils.security import resolve_inside_target
    
    file_path = resolve_inside_target(target_dir, relative_file_path)
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail=f"文件不存在：{relative_file_path}")
    
    if not file_path.is_file():
        raise HTTPException(status_code=400, detail=f"不是文件：{relative_file_path}")
    
    preview = get_file_preview(file_path)
    return {
        "status": "success",
        "preview": preview
    }


async def lock_folder_impl(request) -> dict:
    """锁定目标目录并扫描文件。
    
    Args:
        request: FolderRequest 请求对象
        
    Returns:
        包含文件列表和分析结果的字典
    """
    from models.schemas import FolderRequest
    from utils.security import get_target_dir
    
    if not isinstance(request, FolderRequest):
        raise ValueError("Invalid request type")
    
    target_dir = get_target_dir(request.target_path)
    files_info = scan_target_files(target_dir)
    return {
        "status": "success",
        "target_path": str(target_dir),
        "files": files_info,
        "analysis": build_analysis(files_info),
    }
